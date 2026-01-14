import os, tempfile, subprocess, shutil
from typing import Tuple
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO

def convert_to_xlsx(input_bytes: bytes, filename: str) -> Tuple[bytes, str]:
    # If already xlsx, return
    if filename.lower().endswith(".xlsx"):
        return input_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Convert .ods to .xlsx using LibreOffice
    with tempfile.TemporaryDirectory() as td:
        inp = os.path.join(td, filename)
        with open(inp, "wb") as f:
            f.write(input_bytes)

        subprocess.check_call([
            "soffice", "--headless", "--nologo", "--nolockcheck", "--nodefault", "--norestore",
            "--convert-to", "xlsx", "--outdir", td, inp
        ])
        out = os.path.join(td, os.path.splitext(os.path.basename(filename))[0] + ".xlsx")
        with open(out, "rb") as f:
            data = f.read()
        return data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def fill_values(template_xlsx: bytes, values: list[dict]) -> bytes:
    # values: [{sheet_name, cell_ref, value}]
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "workbook.xlsx")
        with open(path, "wb") as f:
            f.write(template_xlsx)

        wb = load_workbook(path)
        for item in values:
            ws = wb[item["sheet_name"]] if item["sheet_name"] in wb.sheetnames else wb.active
            ws[item["cell_ref"]] = item["value"]
        wb.save(path)
        with open(path, "rb") as f:
            return f.read()

def xlsx_to_pdf(xlsx_bytes: bytes) -> bytes:
    with tempfile.TemporaryDirectory() as td:
        xlsx_path = os.path.join(td, "filled.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_bytes)

        subprocess.check_call([
            "soffice", "--headless", "--nologo", "--nolockcheck", "--nodefault", "--norestore",
            "--convert-to", "pdf", "--outdir", td, xlsx_path
        ])
        pdf_path = os.path.join(td, "filled.pdf")
        with open(pdf_path, "rb") as f:
            return f.read()

def build_audit_pdf(audit_rows: list[dict]) -> bytes:
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    y = h - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Trilha de Auditoria")
    y -= 25
    c.setFont("Helvetica", 9)
    headers = ["Data/Hora (UTC)", "Usuário", "Evento", "Planilha", "Célula", "De", "Para"]
    c.drawString(50, y, " | ".join(headers))
    y -= 15
    c.line(50, y, w-50, y)
    y -= 15

    for r in audit_rows:
        line = f'{r.get("created_at","")} | {r.get("user_email","")} | {r.get("event_type","")} | {r.get("sheet_name","")} | {r.get("cell_ref","")} | {str(r.get("old_value",""))[:25]} | {str(r.get("new_value",""))[:25]}'
        c.drawString(50, y, line)
        y -= 12
        if y < 60:
            c.showPage()
            y = h - 50
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, y, "Trilha de Auditoria (continuação)")
            y -= 25
            c.setFont("Helvetica", 9)

    c.showPage()
    c.save()
    return buf.getvalue()

def merge_pdf_with_audit(pdf_main: bytes, pdf_audit: bytes) -> bytes:
    r1 = PdfReader(BytesIO(pdf_main))
    r2 = PdfReader(BytesIO(pdf_audit))
    w = PdfWriter()
    for p in r1.pages:
        w.add_page(p)
    for p in r2.pages:
        w.add_page(p)
    out = BytesIO()
    w.write(out)
    return out.getvalue()
